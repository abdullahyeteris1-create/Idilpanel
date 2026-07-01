"""Service layer for text library operations."""

from __future__ import annotations

from io import BytesIO
from collections.abc import Mapping
from datetime import datetime
from pathlib import Path
import traceback
from typing import Any

from openpyxl import Workbook, load_workbook

from repositories.text_repository import TextRepository

from .base_service import BaseService


class TextService(BaseService):
    """Validate and orchestrate text library CRUD operations."""

    EXPORT_HEADERS = ["Kur", "Metin Adı", "Kelime Sayısı", "Kategori", "Durum"]

    def __init__(
        self,
        text_repository: TextRepository | None = None,
        repositories: dict[str, Any] | None = None,
    ) -> None:
        merged_repositories = dict(repositories or {})
        if text_repository is not None:
            merged_repositories["text"] = text_repository
        super().__init__(merged_repositories)

    def create_text(self, data):
        return self.get_repository("text").create(self.validate_text(data))

    def get_text(self, record_id):
        return self.get_repository("text").get_by_id(record_id)

    def list_texts(self, limit: int = 500, offset: int = 0):
        return self.get_repository("text").list_all(limit, offset)

    def search_texts(self, query: str = "", course_level: int | None = None, limit: int = 500, offset: int = 0):
        return self.get_repository("text").search(query=query, course_level=course_level, limit=limit, offset=offset)

    def update_text(self, record_id, data):
        return self.get_repository("text").update(record_id, self.validate_text(data))

    def delete_text(self, record_id):
        return self.get_repository("text").delete(record_id)

    def import_texts_from_xlsx(self, file_path: str) -> dict[str, int]:
        path = Path(file_path)
        self._debug_import_log(f"START file_path={path}")
        self._debug_import_log(f"PATH exists={path.exists()} is_file={path.is_file()} suffix={path.suffix}")
        if path.suffix.casefold() != ".xlsx":
            raise ValueError("Sadece .xlsx dosyasi ice aktarilabilir.")

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            self._debug_import_log("OPENPYXL opened workbook")
        except Exception as exc:
            self._debug_import_log("OPENPYXL failed", exc)
            raise ValueError("Excel dosyasi okunamadi.") from exc

        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            self._debug_import_log(f"ROWS loaded count={len(rows)}")
        finally:
            workbook.close()
        if not rows:
            raise ValueError("Excel dosyasi bos.")

        header_map = self._build_header_map(rows[0])
        self._debug_import_log(f"HEADERS raw={rows[0]!r}")
        self._debug_import_log(f"HEADERS mapped={header_map!r}")
        if "course_level" not in header_map:
            raise ValueError("Kur sutunu bulunamadi.")
        if "title" not in header_map:
            raise ValueError("Metin Adi sutunu eksik.")

        processed = 0
        inserted = 0
        skipped = 0

        for row in rows[1:]:
            if not row or all(value is None or str(value).strip() == "" for value in row):
                continue

            processed += 1
            try:
                raw_payload = self._payload_from_excel_row(row, header_map)
                validated = self.validate_text(raw_payload)
            except Exception as exc:
                self._debug_import_log(f"ROW validation failed processed={processed} row={row!r}", exc)
                raise

            try:
                existing = self.get_repository("text").find_by_level_title(
                    validated["course_level"],
                    validated["title"],
                )
            except Exception as exc:
                self._debug_import_log(f"SQL duplicate check failed processed={processed} payload={validated!r}", exc)
                raise
            if existing:
                skipped += 1
                continue

            try:
                self.get_repository("text").create(validated)
            except Exception as exc:
                self._debug_import_log(f"SQL insert failed processed={processed} payload={validated!r}", exc)
                raise
            inserted += 1

        self._debug_import_log(f"DONE processed={processed} inserted={inserted} skipped={skipped}")
        return {
            "processed": processed,
            "inserted": inserted,
            "skipped": skipped,
        }

    def export_texts_xlsx_bytes(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Metin Kutuphanesi"
        sheet.append(self.EXPORT_HEADERS)

        for record in self.search_texts(limit=10000, offset=0):
            sheet.append(
                [
                    f"{record.get('course_level')}. Kur",
                    record.get("title") or "",
                    record.get("word_count") or "",
                    record.get("category") or "",
                    "Aktif" if int(record.get("is_active") or 0) == 1 else "Pasif",
                ]
            )

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def validate_text(self, data: Any | None = None) -> dict[str, Any]:
        if not isinstance(data, Mapping):
            raise ValueError("text data must be a mapping")

        title = str(data.get("title") or data.get("metin_adi") or "").strip()
        if not title:
            raise ValueError("text title is required")

        course_level = self._parse_positive_int(data.get("course_level") or data.get("kur"), "course level")
        category = str(data.get("category") or "").strip() or None
        word_count = self._parse_optional_int(data.get("word_count") or data.get("kelime_sayisi"), "word count")
        is_active = self._parse_active(data.get("is_active", data.get("durum", 1)))

        return {
            "title": title,
            "course_level": course_level,
            "category": category,
            "word_count": word_count,
            "is_active": is_active,
        }

    @classmethod
    def _build_header_map(cls, header_row: tuple[Any, ...]) -> dict[str, int]:
        aliases = {
            "kur": "course_level",
            "courselevel": "course_level",
            "course_level": "course_level",
            "metinadi": "title",
            "metinadı": "title",
            "metinad?": "title",
            "title": "title",
            "baslik": "title",
            "başlık": "title",
            "kelimesayisi": "word_count",
            "kelimesayısı": "word_count",
            "kelimesay?s?": "word_count",
            "wordcount": "word_count",
            "word_count": "word_count",
            "kategori": "category",
            "category": "category",
            "durum": "is_active",
            "status": "is_active",
            "isactive": "is_active",
            "is_active": "is_active",
        }
        result: dict[str, int] = {}
        for index, value in enumerate(header_row):
            normalized = cls._normalize_header(value)
            field_name = aliases.get(normalized)
            if field_name:
                result[field_name] = index
        return result

    @staticmethod
    def _normalize_header(value: Any) -> str:
        text = str(value or "").strip().casefold()
        for char in (" ", "-", ".", "/", "\\", "'", "\""):
            text = text.replace(char, "")
        return text

    def _payload_from_excel_row(self, row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
        def value(field_name: str) -> Any:
            index = header_map.get(field_name)
            if index is None or index >= len(row):
                return ""
            return row[index]

        return {
            "course_level": self._parse_course_level(value("course_level")),
            "title": value("title"),
            "word_count": value("word_count"),
            "category": value("category"),
            "is_active": value("is_active") if value("is_active") not in (None, "") else 1,
        }

    @staticmethod
    def _parse_course_level(value: Any) -> int:
        text = str(value or "").strip()
        if not text:
            raise ValueError("course level is required")
        if text.casefold() == "genel":
            return 1
        digits = "".join(char for char in text if char.isdigit())
        if not digits:
            raise ValueError("course level is required")
        return int(digits)

    @staticmethod
    def _parse_positive_int(value: Any, field_name: str) -> int:
        try:
            parsed = int(str(value).strip())
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{field_name} is required") from exc
        if parsed <= 0:
            raise ValueError(f"{field_name} must be positive")
        return parsed

    @staticmethod
    def _parse_optional_int(value: Any, field_name: str) -> int | None:
        if value is None or str(value).strip() == "":
            return None
        try:
            parsed = int(str(value).strip())
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{field_name} must be a valid integer") from exc
        if parsed < 0:
            raise ValueError(f"{field_name} must not be negative")
        return parsed

    @staticmethod
    def _parse_active(value: Any) -> int:
        text = str(value).strip().casefold()
        if text in {"0", "false", "pasif", "passive", "inactive"}:
            return 0
        return 1

    @staticmethod
    def _debug_import_log(message: str, exc: Exception | None = None) -> None:
        with open("debug_text_library_import.log", "a", encoding="utf-8") as handle:
            handle.write(f"[{datetime.now().isoformat(timespec='seconds')}] {message}\n")
            if exc is not None:
                handle.write(f"exception_type: {type(exc).__name__}\n")
                handle.write(f"exception_message: {exc}\n")
                handle.write(traceback.format_exc())
                handle.write("\n")
