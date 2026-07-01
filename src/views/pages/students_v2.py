"""LEGACY Students V2 page.

Canonical Students page: views.pages.students_v3.build_students_v3_page.
This module is retained for reference/import-export helpers only and must not
receive new feature work.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path

import flet as ft
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from components import (
    PageContainer,
    build_action_panel,
    build_app_card,
    build_app_datepicker,
    build_app_dropdown,
    build_app_header,
    build_badge,
    build_dialog,
    build_danger_button,
    build_empty_state,
    build_error_state,
    build_filter_bar,
    build_form_card,
    build_loading_state,
    build_primary_button,
    build_search_bar,
    build_secondary_button,
    build_table_card,
    build_text_field,
)
from controllers import build_course_controller, build_student_controller
from controllers.course_controller import CourseController
from controllers.student_controller import StudentController


PHONE_PATTERN = re.compile(r"^0?\d{10}$")
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
GRADE_PATTERN = re.compile(r"(\d+)")
IMPORT_HEADERS = {
    "ad soyad": "ad_soyad",
    "ad_soyad": "ad_soyad",
    "sinif": "sinif",
    "kur": "kur",
    "telefon": "telefon",
    "e-posta": "email",
    "eposta": "email",
    "email": "email",
    "kullanici adi": "kullanici_adi",
    "kullanici_adi": "kullanici_adi",
    "sifre": "sifre",
    "baslangic": "baslangic_tarihi",
    "baslangic tarihi": "baslangic_tarihi",
    "baslangic_tarihi": "baslangic_tarihi",
    "bitis": "bitis_tarihi",
    "bitis tarihi": "bitis_tarihi",
    "bitis_tarihi": "bitis_tarihi",
    "durum": "durum",
    "veli adi": "veli_adi",
    "veli_adi": "veli_adi",
    "notlar": "notlar",
}
EXPORT_COLUMNS = ["Ad Soyad", "Sınıf", "Kur", "Telefon", "Durum", "Başlangıç"]

IS_LEGACY_STUDENTS_PAGE = True
CANONICAL_STUDENTS_MODULE = "views.pages.students_v3"


def _status_variant(status: str) -> str:
    normalized = status.strip().lower()
    if normalized == "aktif":
        return "success"
    if normalized == "tamamlandi":
        return "primary"
    return "warning"


def _friendly_error(exc: Exception) -> str:
    message = str(exc)
    if "student name cannot be empty" in message:
        return "Ad Soyad alanı zorunludur."
    if "student class information cannot be empty" in message:
        return "Sınıf alanı zorunludur."
    if "student start date cannot be empty" in message:
        return "Başlangıç tarihi seçiniz."
    if "student start date must be a valid date" in message:
        return "Başlangıç tarihi geçersizdir."
    if "student end date must be a valid date" in message:
        return "Bitiş tarihi geçersizdir."
    if "student end date cannot be before start date" in message:
        return "Bitiş tarihi, başlangıç tarihinden önce olamaz."
    if "student email must be valid" in message:
        return "E-posta formatı geçersizdir."
    if "student status must be Aktif or Beklemede" in message:
        return "Durum alanı geçersizdir."
    if "no such table" in message:
        return "Veritabanı hazır değil. Lütfen kurulum adımlarını kontrol edin."
    if "UNIQUE constraint failed: students.kullanici_adi" in message:
        return "Kullanıcı adı benzersiz olmalıdır."
    if "course name cannot be empty" in message:
        return "Kurs seçimi zorunludur."
    if "an active course with the same name already exists" in message:
        return "Aynı kurs ikinci kez atanamaz."
    if "UNIQUE constraint failed: courses.student_id, courses.kur_no, courses.is_active" in message:
        return "Aynı kurs ikinci kez atanamaz."
    return "Islem su anda tamamlanamadi. Lutfen tekrar deneyin."


def _parse_iso_date(value: str) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _extract_grade(value: str) -> int | None:
    match = GRADE_PATTERN.search(value or "")
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _extract_kur_value_from_notes(note_text: str) -> int | None:
    for line in (note_text or "").splitlines():
        if not line.startswith("Kur:"):
            continue
        kur_grade = _extract_grade(line)
        if kur_grade is not None:
            return kur_grade
    return None


def _status_key(status: str) -> str:
    normalized = (status or "").strip().lower()
    if normalized == "aktif":
        return "Aktif"
    if normalized in {"pasif", "beklemede"}:
        return "Pasif"
    if normalized in {"tamamlandi", "egitimi tamamlandi"}:
        return "Egitimi Tamamlandi"
    return "Pasif"


def _students_v2_apply_filters(
    records: list[dict],
    search_query: str,
    status_filter: str,
    class_filter: str,
    kur_filter: str,
    start_from: str,
    start_to: str,
    sort_field: str,
    sort_direction: str,
) -> list[dict]:
    query = (search_query or "").strip().lower()
    status_value = (status_filter or "Tumu").strip()
    class_value = (class_filter or "Tumu").strip()
    kur_value = (kur_filter or "Tumu").strip()
    date_from = _parse_iso_date(start_from)
    date_to = _parse_iso_date(start_to)

    filtered = list(records)

    if query:
        def _matches_query(item: dict) -> bool:
            haystack = [
                str(item.get("ad_soyad") or ""),
                str(item.get("veli_adi") or ""),
                str(item.get("telefon") or ""),
                str(item.get("email") or item.get("eposta") or ""),
                str(item.get("kullanici_adi") or ""),
            ]
            return query in " ".join(haystack).lower()

        filtered = [item for item in filtered if _matches_query(item)]

    if status_value != "Tumu":
        filtered = [item for item in filtered if _status_key(str(item.get("durum") or "")) == status_value]

    if class_value != "Tumu":
        class_grade = _extract_grade(class_value)
        if class_grade is not None:
            filtered = [
                item
                for item in filtered
                if _extract_grade(str(item.get("sinif") or "")) == class_grade
            ]

    if kur_value != "Tumu":
        kur_grade = _extract_grade(kur_value)
        if kur_grade is not None:
            filtered = [
                item
                for item in filtered
                if _extract_kur_value_from_notes(str(item.get("notlar") or "")) == kur_grade
            ]

    if date_from or date_to:
        def _date_in_range(item: dict) -> bool:
            start_date = _parse_iso_date(str(item.get("baslangic_tarihi") or ""))
            if start_date is None:
                return False
            if date_from and start_date < date_from:
                return False
            if date_to and start_date > date_to:
                return False
            return True

        filtered = [item for item in filtered if _date_in_range(item)]

    reverse = (sort_direction or "Artan") == "Azalan"

    def _sort_key(item: dict):
        if sort_field == "Baslangic Tarihi":
            return _parse_iso_date(str(item.get("baslangic_tarihi") or "")) or date.min
        if sort_field == "Kur":
            return _extract_kur_value_from_notes(str(item.get("notlar") or "")) or 0
        if sort_field == "Sinif":
            return _extract_grade(str(item.get("sinif") or "")) or 0
        return str(item.get("ad_soyad") or "").lower()

    filtered.sort(key=_sort_key, reverse=reverse)
    return filtered


def students_v2_responsive_profile(width: int) -> str:
    """Classify viewport width using Students V2 target breakpoints."""

    if width >= 1800:
        return "1920 px"
    if width >= 1500:
        return "1600 px"
    if width >= 1320:
        return "1366 px"
    if width >= 1200:
        return "1280 px"
    if width >= 768:
        return "Tablet"
    return "Mobil"


def students_v2_tab_order() -> list[str]:
    """Return the intended tab order for focus and keyboard workflow checks."""

    return [
        "Ad Soyad",
        "Sinif",
        "Veli Adi",
        "Telefon",
        "E-posta",
        "Kullanici Adi",
        "Sifre",
        "Baslangic Tarihi",
        "Bitis Tarihi",
        "Kur",
        "Durum",
        "Notlar",
    ]


def _safe_text(value: object) -> str:
    return str(value or "").strip()


def students_v2_validate_course_assignment(
    student_id: int | None,
    selected_kur: str,
    existing_courses: list[dict],
) -> tuple[bool, str]:
    if not int(student_id or 0):
        return False, "Öğrenci seçili olmalıdır."

    if not _safe_text(selected_kur):
        return False, "Kurs seçili olmalıdır."

    kur_value = _extract_grade(selected_kur)
    if kur_value is None:
        return False, "Kurs seçimi geçersizdir."

    for course in existing_courses:
        if int(course.get("student_id", 0) or 0) != int(student_id):
            continue
        if int(course.get("kur_no", 0) or 0) != int(kur_value):
            continue

        if str(course.get("durum") or "").strip() != "Aktif":
            return False, "Pasif kurs seçilemez."
        return False, "Aynı kurs ikinci kez atanamaz."

    return True, ""


def students_v2_active_course_label(student_id: int, courses: list[dict]) -> str:
    active_rows = [
        item
        for item in courses
        if int(item.get("student_id", 0) or 0) == int(student_id)
        and str(item.get("durum") or "").strip() == "Aktif"
    ]
    if not active_rows:
        return "-"

    active_rows.sort(key=lambda row: int(row.get("id", 0) or 0), reverse=True)
    latest = active_rows[0]
    kur_no = int(latest.get("kur_no", 0) or 0)
    return f"Kur {kur_no}" if kur_no > 0 else "-"


def _normalize_import_header(value: object) -> str:
    return _safe_text(value).lower()


def _student_row_to_export_row(record: dict) -> list[str]:
    kur_value = _extract_kur_value_from_notes(_safe_text(record.get("notlar")))
    return [
        _safe_text(record.get("ad_soyad")),
        _safe_text(record.get("sinif")),
        str(kur_value) if kur_value is not None else "",
        _safe_text(record.get("telefon")),
        _safe_text(record.get("durum") or "Aktif"),
        _safe_text(record.get("baslangic_tarihi")),
    ]


def students_v2_export_excel_bytes(records: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ogrenciler"
    sheet.append(EXPORT_COLUMNS)
    for record in records:
        sheet.append(_student_row_to_export_row(record))

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def students_v2_export_csv_text(records: list[dict]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_COLUMNS)
    for record in records:
        writer.writerow(_student_row_to_export_row(record))
    return output.getvalue()


def students_v2_export_pdf_bytes(records: list[dict], filter_info: str) -> bytes:
    output = io.BytesIO()
    document = SimpleDocTemplate(output, pagesize=A4)
    styles = getSampleStyleSheet()

    elements: list = [
        Paragraph("IDIL HIZLI OKUMA - Ogrenci Raporu", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
        Paragraph(f"Toplam kayit: {len(records)}", styles["Normal"]),
        Paragraph(f"Filtre: {filter_info or 'Yok'}", styles["Normal"]),
        Spacer(1, 10),
    ]

    table_data = [EXPORT_COLUMNS]
    for record in records:
        table_data.append(_student_row_to_export_row(record))

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5E7EB")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#9CA3AF")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    document.build(elements)
    return output.getvalue()


def students_v2_parse_import_xlsx(file_path: str, existing_usernames: set[str]) -> dict:
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return {"total": 0, "valid_rows": [], "errors": ["Dosya bos."]}

    header_map: dict[int, str] = {}
    for index, cell in enumerate(rows[0]):
        normalized = _normalize_import_header(cell)
        if normalized in IMPORT_HEADERS:
            header_map[index] = IMPORT_HEADERS[normalized]

    if not header_map:
        return {"total": 0, "valid_rows": [], "errors": ["Gecerli kolon bulunamadi."]}

    valid_rows: list[dict] = []
    errors: list[str] = []
    seen_usernames: set[str] = set()

    for row_number, raw_row in enumerate(rows[1:], start=2):
        if raw_row is None:
            continue

        parsed: dict[str, str] = {}
        for column_index, key in header_map.items():
            parsed[key] = _safe_text(raw_row[column_index] if column_index < len(raw_row) else "")

        if not any(parsed.values()):
            continue

        row_errors: list[str] = []
        if not parsed.get("ad_soyad"):
            row_errors.append("Ad Soyad zorunludur")
        if not parsed.get("sinif"):
            row_errors.append("Sinif zorunludur")
        if not parsed.get("kur"):
            row_errors.append("Kur zorunludur")

        phone = parsed.get("telefon", "")
        if phone and not PHONE_PATTERN.fullmatch(re.sub(r"\s+", "", phone)):
            row_errors.append("Telefon formatı geçersizdir")

        email = parsed.get("email", "")
        if email and not EMAIL_PATTERN.fullmatch(email):
            row_errors.append("E-posta formatı geçersizdir")

        username = parsed.get("kullanici_adi", "").lower()
        if username:
            if username in existing_usernames or username in seen_usernames:
                row_errors.append("Aynı kullanıcı adı")
            else:
                seen_usernames.add(username)

        if row_errors:
            errors.append(f"Satir {row_number}: {', '.join(row_errors)}")
            continue

        notes = _safe_text(parsed.get("notlar"))
        kur_text = _safe_text(parsed.get("kur"))
        if notes:
            notes = f"{notes}\nKur: {kur_text}"
        else:
            notes = f"Kur: {kur_text}"

        valid_rows.append(
            {
                "ad_soyad": _safe_text(parsed.get("ad_soyad")),
                "sinif": _safe_text(parsed.get("sinif")),
                "veli_adi": _safe_text(parsed.get("veli_adi")),
                "telefon": phone,
                "email": email,
                "kullanici_adi": _safe_text(parsed.get("kullanici_adi")),
                "sifre": _safe_text(parsed.get("sifre")),
                "baslangic_tarihi": _safe_text(parsed.get("baslangic_tarihi")) or date.today().isoformat(),
                "bitis_tarihi": _safe_text(parsed.get("bitis_tarihi")),
                "durum": _safe_text(parsed.get("durum")) or "Aktif",
                "notlar": notes,
            }
        )

    return {"total": max(0, len(rows) - 1), "valid_rows": valid_rows, "errors": errors}


def build_students_v2_page() -> ft.Control:
    """Build Students V2 page with complete daily teacher workflow UX."""

    controller: StudentController = build_student_controller()
    course_controller: CourseController = build_course_controller()
    state = {
        "students": [],
        "courses": [],
        "selected_id": None,
        "selected_snapshot": None,
        "view_state": "loading",
        "page": 1,
        "page_size": 6,
        "search_query": "",
        "status_filter": "Tumu",
        "class_filter": "Tumu",
        "kur_filter": "Tumu",
        "start_from": "",
        "start_to": "",
        "sort_field": "Ad Soyad",
        "sort_direction": "Artan",
        "last_filtered_rows": [],
        "import_preview_rows": [],
        "import_errors": [],
        "pending_picker_action": None,
    }

    table_host = ft.Column(spacing=10)

    feedback_tag_host = ft.Container(content=build_badge(text="Bilgi", variant="primary"))
    feedback_text = ft.Text("Hazır")
    feedback_row = ft.Row(spacing=8, controls=[feedback_tag_host, feedback_text])
    result_text = ft.Text("", weight=ft.FontWeight.W_600)
    active_filters_row = ft.Row(spacing=8, wrap=True)

    name_field = build_text_field("Ad Soyad", required=True)
    class_field = build_text_field("Sinif", required=True)
    parent_field = build_text_field("Veli Adi")
    phone_field = build_text_field("Telefon")
    email_field = build_text_field("E-posta")
    username_field = build_text_field("Kullanici Adi")
    password_field = build_text_field("Sifre", password=True)
    start_date = build_app_datepicker("Baslangic Tarihi", required=True)
    end_date = build_app_datepicker("Bitis Tarihi")
    kur_dropdown = build_app_dropdown("Kur", options=[str(i) for i in range(1, 17)])
    status_dropdown = build_app_dropdown("Durum", options=["Aktif", "Beklemede"], value="Aktif")
    notes_field = build_text_field("Notlar", multiline=True, min_lines=4, max_lines=8)

    search_bar = build_search_bar(
        title="Arama",
        subtitle="Ad Soyad, Veli Adi, Telefon, E-posta, Kullanici Adi",
        hint_text="Ogrenci ara...",
    )

    filter_status = build_app_dropdown(
        label="Durum",
        options=["Tumu", "Aktif", "Pasif", "Egitimi Tamamlandi"],
        value="Tumu",
        on_change=lambda e: _handle_filter_change(e.page),
    )
    filter_kur = build_app_dropdown(
        label="Kur",
        options=["Tumu"] + [f"Kur {i}" for i in range(1, 13)],
        value="Tumu",
        on_change=lambda e: _handle_filter_change(e.page),
    )
    filter_class = build_app_dropdown(
        label="Sinif",
        options=["Tumu"] + [f"{i}. Sinif" for i in range(1, 13)],
        value="Tumu",
        on_change=lambda e: _handle_filter_change(e.page),
    )
    filter_start_from = build_app_datepicker(
        "Baslangic",
        on_date_change=lambda _: _handle_filter_change(None),
    )
    filter_start_to = build_app_datepicker(
        "Bitis",
        on_date_change=lambda _: _handle_filter_change(None),
    )
    sort_field_dropdown = build_app_dropdown(
        label="Siralama",
        options=["Ad Soyad", "Baslangic Tarihi", "Kur", "Sinif"],
        value="Ad Soyad",
        on_change=lambda e: _handle_filter_change(e.page),
    )
    sort_direction_dropdown = build_app_dropdown(
        label="Yon",
        options=["Artan", "Azalan"],
        value="Artan",
        on_change=lambda e: _handle_filter_change(e.page),
    )

    def _set_feedback(kind: str, message: str) -> None:
        variant = {
            "success": "success",
            "info": "primary",
            "warning": "warning",
            "error": "danger",
        }.get(kind, "primary")
        label = {
            "success": "Basari",
            "info": "Bilgi",
            "warning": "Uyari",
            "error": "Hata",
        }.get(kind, "Bilgi")
        feedback_tag_host.content = build_badge(text=label, variant=variant)
        feedback_text.value = message

    def _active_filter_text() -> str:
        parts: list[str] = []
        if state["search_query"]:
            parts.append(f"Arama={state['search_query']}")
        if state["status_filter"] != "Tumu":
            parts.append(f"Durum={state['status_filter']}")
        if state["class_filter"] != "Tumu":
            parts.append(f"Sinif={state['class_filter']}")
        if state["kur_filter"] != "Tumu":
            parts.append(f"Kur={state['kur_filter']}")
        if state["start_from"]:
            parts.append(f"Baslangic>={state['start_from']}")
        if state["start_to"]:
            parts.append(f"Baslangic<={state['start_to']}")
        return ", ".join(parts)

    def _write_export_file(directory: str, file_name: str, content: bytes | str) -> str:
        export_dir = Path(directory)
        export_dir.mkdir(parents=True, exist_ok=True)
        file_path = export_dir / file_name
        if isinstance(content, str):
            file_path.write_text(content, encoding="utf-8")
        else:
            file_path.write_bytes(content)
        return str(file_path)

    def _load_courses() -> None:
        # Temporarily remove try/except to expose any exceptions
        state["courses"] = list(course_controller.list_courses(limit=1000, offset=0))

    def _student_courses(student_id: int) -> list[dict]:
        return [
            item
            for item in state["courses"]
            if int(item.get("student_id", 0) or 0) == int(student_id)
        ]

    def _open_course_assignment_modal(record: dict, page: ft.Page | None) -> None:
        if page is None:
            _set_feedback("warning", "Sayfa hazır değil.")
            return

        student_id = int(record.get("id", 0) or 0)
        if not student_id:
            _set_feedback("warning", "Öğrenci seçili olmalıdır.")
            page.update()
            return

        existing_rows = _student_courses(student_id)
        assigned_kur_values = {int(item.get("kur_no", 0) or 0) for item in existing_rows}
        available_kur_options = [
            f"Kur {value}"
            for value in range(1, 13)
            if value not in assigned_kur_values
        ]

        selected_course_dropdown = build_app_dropdown(
            label="Kurs Secimi",
            options=available_kur_options or ["Uygun kurs yok"],
            value=None,
        )
        selected_course_dropdown.disabled = not available_kur_options

        assigned_text = [
            f"Kur {int(item.get('kur_no', 0) or 0)} - {str(item.get('durum') or '-') }"
            for item in sorted(existing_rows, key=lambda row: int(row.get("kur_no", 0) or 0))
        ]
        assigned_card = build_app_card(
            title="Atanan Kurslar",
            subtitle="Ogrenciye atanmis kurslar",
            content=ft.Column(
                spacing=6,
                controls=[ft.Text(value) for value in (assigned_text or ["Henuz kurs atanmamis."])],
            ),
        )

        modal = ft.AlertDialog(
            modal=True,
            title=ft.Text(f"Kurs Ata - {str(record.get('ad_soyad') or '')}"),
            content=ft.Container(
                width=560,
                content=ft.Column(
                    spacing=12,
                    tight=True,
                    controls=[selected_course_dropdown, assigned_card],
                ),
            ),
            actions_alignment=ft.MainAxisAlignment.END,
        )

        def _close_dialog(_: ft.ControlEvent | None = None) -> None:
            page.close(modal)
            page.update()

        def _save_assignment(e: ft.ControlEvent) -> None:
            selected_kur_text = str(selected_course_dropdown.value or "").strip()
            valid, message = students_v2_validate_course_assignment(
                student_id=student_id,
                selected_kur=selected_kur_text,
                existing_courses=existing_rows,
            )
            if not valid:
                _set_feedback("warning", message)
                e.page.update()
                return

            kur_value = _extract_grade(selected_kur_text)
            if kur_value is None:
                _set_feedback("warning", "Kurs secimi gecersiz.")
                e.page.update()
                return

            try:
                course_controller.assign_course_to_student(student_id, kur_value)
                _load_courses()
                _render_table(e.page)
                _set_feedback("success", f"Kurs atamasi tamamlandi: Kur {kur_value}")
                _close_dialog(e)
            except Exception as exc:
                _set_feedback("error", f"Kurs atamasi basarisiz: {_friendly_error(exc)}")
                e.page.update()

        modal.actions = [
            build_secondary_button("Vazgec", on_click=_close_dialog),
            build_primary_button("Kaydet", on_click=_save_assignment, disabled=not available_kur_options),
        ]

        page.open(modal)
        page.update()

    def _export_records(page: ft.Page | None, fmt: str, directory: str) -> None:
        records = list(state.get("last_filtered_rows") or [])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filter_info = _active_filter_text()

        if fmt == "xlsx":
            data = students_v2_export_excel_bytes(records)
            file_path = _write_export_file(directory, f"students_v2_{timestamp}.xlsx", data)
        elif fmt == "pdf":
            data = students_v2_export_pdf_bytes(records, filter_info)
            file_path = _write_export_file(directory, f"students_v2_{timestamp}.pdf", data)
        else:
            data = students_v2_export_csv_text(records)
            file_path = _write_export_file(directory, f"students_v2_{timestamp}.csv", data)

        _set_feedback("success", f"Dışa aktarma tamamlandı. Dosya: {file_path}")
        if page:
            page.update()

    def _handle_import_confirm(page: ft.Page | None) -> None:
        valid_rows = list(state.get("import_preview_rows") or [])
        validation_errors = list(state.get("import_errors") or [])
        added_count = 0
        runtime_errors: list[str] = []

        for row_index, payload in enumerate(valid_rows, start=1):
            try:
                controller.create_student(payload)
                added_count += 1
            except Exception as exc:
                runtime_errors.append(f"Satir {row_index}: {_friendly_error(exc)}")

        total_errors = len(validation_errors) + len(runtime_errors)
        _load_students()
        _render_table(page)

        summary = (
            "İçe aktarma tamamlandı.\n"
            f"{added_count} kayıt eklendi.\n"
            f"{total_errors} kayıt hata nedeniyle eklenemedi."
        )
        _set_feedback("success" if total_errors == 0 else "warning", summary)

        all_errors = validation_errors + runtime_errors
        if all_errors and page is not None:
            preview_errors = "\n".join(all_errors[:10])
            dialog = build_dialog(
                title="Hata Raporu",
                content=preview_errors,
                confirm_text="Tamam",
                cancel_text="Kapat",
                on_confirm=lambda e: (e.page.close(dialog), e.page.update()),
                on_cancel=lambda e: (e.page.close(dialog), e.page.update()),
            )
            page.open(dialog)

        if page:
            page.update()

    def _handle_file_picker_result(e: ft.FilePickerResultEvent) -> None:
        action = state.get("pending_picker_action")
        state["pending_picker_action"] = None

        if action == "import":
            files = getattr(e, "files", None) or []
            if not files:
                _set_feedback("info", "İçe aktarma iptal edildi.")
                e.page.update()
                return

            selected_file = files[0]
            try:
                existing_usernames = {
                    _safe_text(item.get("kullanici_adi")).lower()
                    for item in state["students"]
                    if _safe_text(item.get("kullanici_adi"))
                }
                parsed = students_v2_parse_import_xlsx(selected_file.path, existing_usernames)
            except Exception as exc:
                _set_feedback("error", f"Bozuk dosya: {_friendly_error(exc)}")
                e.page.update()
                return

            state["import_preview_rows"] = list(parsed["valid_rows"])
            state["import_errors"] = list(parsed["errors"])

            if parsed["total"] == 0:
                _set_feedback("warning", "Boş dosya tespit edildi.")
                e.page.update()
                return

            preview_text = (
                f"Toplam satir: {parsed['total']}\n"
                f"Gecerli satir: {len(parsed['valid_rows'])}\n"
                f"Hatali satir: {len(parsed['errors'])}"
            )

            def _confirm_import(event: ft.ControlEvent) -> None:
                event.page.close(preview_dialog)
                _handle_import_confirm(event.page)

            def _cancel_import(event: ft.ControlEvent) -> None:
                event.page.close(preview_dialog)
                _set_feedback("info", "İçe aktarma iptal edildi.")
                event.page.update()

            preview_dialog = build_dialog(
                title="İçe Aktarma Önizleme",
                content=preview_text,
                confirm_text="İçe Aktar",
                cancel_text="Vazgec",
                on_confirm=_confirm_import,
                on_cancel=_cancel_import,
            )
            e.page.open(preview_dialog)
            e.page.update()
            return

        export_path = getattr(e, "path", None)
        if action in {"export_xlsx", "export_pdf", "export_csv"}:
            if not export_path:
                _set_feedback("info", "Dışa aktarma iptal edildi.")
                e.page.update()
                return
            fmt = action.replace("export_", "")
            _export_records(e.page, fmt, export_path)

    file_picker = ft.FilePicker()
    file_picker.on_result = _handle_file_picker_result

    def _ensure_file_picker(page: ft.Page | None) -> bool:
        if page is None:
            _set_feedback("warning", "Sayfa hazır değil.")
            return False
        if file_picker not in page.overlay:
            page.overlay.append(file_picker)
        return True

    def _handle_import_pick(e: ft.ControlEvent) -> None:
        if not _ensure_file_picker(e.page):
            return
        state["pending_picker_action"] = "import"
        file_picker.pick_files(
            dialog_title="Excel dosyasi sec",
            allowed_extensions=["xlsx"],
            allow_multiple=False,
        )

    def _handle_export_pick(e: ft.ControlEvent, fmt: str) -> None:
        if not _ensure_file_picker(e.page):
            return
        state["pending_picker_action"] = f"export_{fmt}"
        file_picker.get_directory_path(dialog_title="Disa aktarma klasoru sec")

    def _focus_first_field(page: ft.Page | None) -> None:
        if page is None:
            return
        try:
            page.set_focus(name_field)
        except Exception:
            pass

    def _get_date_value(control: ft.Control) -> str:
        field = control.data["field"]
        return str(field.value or "").strip()

    def _set_date_value(control: ft.Control, value: str) -> None:
        field = control.data["field"]
        field.value = value

    def _extract_filter_date(control: ft.Control) -> str:
        field = control.data["field"]
        return str(field.value or "").strip()

    def _set_filter_date(control: ft.Control, value: str) -> None:
        field = control.data["field"]
        field.value = value

    def _sync_filter_state() -> None:
        if isinstance(search_input_field, ft.TextField):
            state["search_query"] = str(search_input_field.value or "").strip()
        state["status_filter"] = str(filter_status.value or "Tumu")
        state["class_filter"] = str(filter_class.value or "Tumu")
        state["kur_filter"] = str(filter_kur.value or "Tumu")
        state["start_from"] = _extract_filter_date(filter_start_from)
        state["start_to"] = _extract_filter_date(filter_start_to)
        state["sort_field"] = str(sort_field_dropdown.value or "Ad Soyad")
        state["sort_direction"] = str(sort_direction_dropdown.value or "Artan")

    def _render_active_filters() -> None:
        chips: list[ft.Control] = []

        if state["status_filter"] != "Tumu":
            chips.append(build_badge(text=f"✓ {state['status_filter']}", variant="primary"))
        if state["kur_filter"] != "Tumu":
            chips.append(build_badge(text=f"✓ {state['kur_filter']}", variant="primary"))
        if state["class_filter"] != "Tumu":
            chips.append(build_badge(text=f"✓ {state['class_filter']}", variant="primary"))
        if state["start_from"]:
            chips.append(build_badge(text=f"✓ Baslangic >= {state['start_from']}", variant="primary"))
        if state["start_to"]:
            chips.append(build_badge(text=f"✓ Baslangic <= {state['start_to']}", variant="primary"))
        if state["search_query"]:
            chips.append(build_badge(text=f"✓ Arama: {state['search_query']}", variant="primary"))

        if chips:
            active_filters_row.controls = [ft.Text("Aktif Filtreler:", weight=ft.FontWeight.W_600)] + chips
        else:
            active_filters_row.controls = [ft.Text("Aktif Filtreler: Yok")]

    def _handle_filter_change(page: ft.Page | None) -> None:
        _sync_filter_state()
        state["page"] = 1
        _render_table(page)
        if page:
            page.update()

    def _handle_clear_filters(e: ft.ControlEvent) -> None:
        state["search_query"] = ""
        state["status_filter"] = "Tumu"
        state["class_filter"] = "Tumu"
        state["kur_filter"] = "Tumu"
        state["start_from"] = ""
        state["start_to"] = ""
        state["sort_field"] = "Ad Soyad"
        state["sort_direction"] = "Artan"

        if isinstance(search_input_field, ft.TextField):
            search_input_field.value = ""
        filter_status.value = "Tumu"
        filter_class.value = "Tumu"
        filter_kur.value = "Tumu"
        _set_filter_date(filter_start_from, "")
        _set_filter_date(filter_start_to, "")
        sort_field_dropdown.value = "Ad Soyad"
        sort_direction_dropdown.value = "Artan"

        state["page"] = 1
        _set_feedback("info", "Filtreler temizlendi.")
        _render_table(e.page)
        e.page.update()

    def _extract_kur(note_text: str) -> str:
        for line in (note_text or "").splitlines():
            if line.startswith("Kur:"):
                return line.split(":", 1)[1].strip()
        return ""

    def _plain_notes(note_text: str) -> str:
        lines: list[str] = []
        for line in (note_text or "").splitlines():
            if line.startswith("Kur:"):
                continue
            lines.append(line)
        return "\n".join(lines).strip()

    def _build_notes() -> str:
        lines: list[str] = []
        plain = str(notes_field.value or "").strip()
        kur_value = str(kur_dropdown.value or "").strip()

        if plain:
            lines.append(plain)
        if kur_value:
            lines.append(f"Kur: {kur_value}")

        return "\n".join(lines)

    def _validate_phone(value: str) -> bool:
        cleaned = re.sub(r"\s+", "", value)
        return bool(PHONE_PATTERN.fullmatch(cleaned))

    def _validate_email(value: str) -> bool:
        return bool(EMAIL_PATTERN.fullmatch(value))

    def _validate_username_unique(username: str) -> bool:
        selected_id = state.get("selected_id")
        for item in state["students"]:
            if int(item.get("id", 0) or 0) == int(selected_id or 0):
                continue
            if str(item.get("kullanici_adi") or "").strip().lower() == username.lower():
                return False
        return True

    def _validate_form() -> str | None:
        if not str(name_field.value or "").strip():
            return "Ad Soyad alanı zorunludur."
        if not str(class_field.value or "").strip():
            return "Sınıf alanı zorunludur."
        if not _get_date_value(start_date):
            return "Başlangıç tarihi zorunludur."

        phone = str(phone_field.value or "").strip()
        if phone and not _validate_phone(phone):
            return "Telefon numarası geçersizdir."

        email = str(email_field.value or "").strip()
        if email and not _validate_email(email):
            return "E-posta formatı geçersizdir."

        username = str(username_field.value or "").strip()
        if username and not _validate_username_unique(username):
            return "Kullanıcı adı benzersiz olmalıdır."

        return None

    def _payload() -> dict[str, str]:
        return {
            "ad_soyad": str(name_field.value or "").strip(),
            "sinif": str(class_field.value or "").strip(),
            "veli_adi": str(parent_field.value or "").strip(),
            "telefon": str(phone_field.value or "").strip(),
            "email": str(email_field.value or "").strip(),
            "kullanici_adi": str(username_field.value or "").strip(),
            "sifre": str(password_field.value or "").strip(),
            "baslangic_tarihi": _get_date_value(start_date),
            "bitis_tarihi": _get_date_value(end_date),
            "durum": str(status_dropdown.value or "Aktif").strip(),
            "notlar": _build_notes(),
        }

    def _clear_form() -> None:
        state["selected_id"] = None
        state["selected_snapshot"] = None
        name_field.value = ""
        class_field.value = ""
        parent_field.value = ""
        phone_field.value = ""
        email_field.value = ""
        username_field.value = ""
        password_field.value = ""
        _set_date_value(start_date, "")
        _set_date_value(end_date, "")
        kur_dropdown.value = None
        notes_field.value = ""
        status_dropdown.value = "Aktif"

    def _fill_form(record: dict) -> None:
        state["selected_id"] = int(record.get("id", 0) or 0)
        state["selected_snapshot"] = dict(record)

        notes_raw = str(record.get("notlar") or "")
        name_field.value = str(record.get("ad_soyad") or "")
        class_field.value = str(record.get("sinif") or "")
        parent_field.value = str(record.get("veli_adi") or "")
        phone_field.value = str(record.get("telefon") or "")
        email_field.value = str(record.get("email") or record.get("eposta") or "")
        username_field.value = str(record.get("kullanici_adi") or "")
        password_field.value = str(record.get("sifre") or "")
        _set_date_value(start_date, str(record.get("baslangic_tarihi") or ""))
        _set_date_value(end_date, str(record.get("bitis_tarihi") or ""))
        kur_dropdown.value = _extract_kur(notes_raw) or None
        notes_field.value = _plain_notes(notes_raw)
        status_dropdown.value = str(record.get("durum") or "Aktif")

    def _load_students() -> None:
        try:
            state["view_state"] = "loading"
            records = list(controller.list_students(limit=1000, offset=0))
            state["students"] = records
            _load_courses()
            state["view_state"] = "ready"
            if state["page"] < 1:
                state["page"] = 1
        except Exception as exc:
            # DEBUG: Log the actual exception
            import traceback
            tb = traceback.format_exc()
            with open('debug_load_students_error.log', 'w', encoding='utf-8') as f:
                f.write(tb)
            
            # Also print to console
            print("\n" + "="*70)
            print("EXCEPTION IN _load_students():")
            print("="*70)
            print(tb)
            print("="*70 + "\n")
            
            state["view_state"] = "error"
            _set_feedback("error", _friendly_error(exc))

    def _refresh_form_from_selection() -> None:
        selected_id = int(state.get("selected_id") or 0)
        if not selected_id:
            return
        for record in state["students"]:
            if int(record.get("id", 0) or 0) == selected_id:
                _fill_form(record)
                return

    def _handle_new(e: ft.ControlEvent) -> None:
        _clear_form()
        _set_feedback("info", "Yeni ogrenci formu acildi.")
        _render_table(e.page)
        _focus_first_field(e.page)
        e.page.update()

    def _handle_clear(e: ft.ControlEvent) -> None:
        _clear_form()
        _set_feedback("info", "Form temizlendi.")
        _render_table(e.page)
        _focus_first_field(e.page)
        e.page.update()

    def _handle_cancel(e: ft.ControlEvent) -> None:
        snapshot = state.get("selected_snapshot")
        if snapshot:
            _fill_form(snapshot)
            _set_feedback("warning", "Değişikliklerden vazgeçildi.")
        else:
            _clear_form()
            _set_feedback("warning", "Değişikliklerden vazgeçildi.")
        _render_table(e.page)
        _focus_first_field(e.page)
        e.page.update()

    def _handle_list(e: ft.ControlEvent) -> None:
        _load_students()
        _render_table(e.page)
        _set_feedback("info", "Öğrenci listesi güncellendi.")
        e.page.update()

    def _handle_create(e: ft.ControlEvent) -> None:
        validation_error = _validate_form()
        if validation_error:
            _set_feedback("warning", validation_error)
            e.page.update()
            return

        payload = _payload()
        
        try:
            result = controller.create_student(payload)
            _load_students()
            _clear_form()
            
            # Wrap table rendering in try/except to catch UI errors
            try:
                _render_table(e.page)
            except Exception as render_error:
                # If table rendering fails, show error but continue
                print(f"Table render error: {render_error}")
                import traceback
                print(traceback.format_exc())
                _set_feedback("warning", "Liste gösterilirken hata oluştu. Lütfen listeyi yenileyin.")
            
            _set_feedback("success", "Öğrenci başarıyla oluşturuldu.")
            _focus_first_field(e.page)
            e.page.update()
        
        except Exception as exc:
            # Log and show error
            print(f"Create student error: {exc}")
            import traceback
            print(traceback.format_exc())
            _set_feedback("error", _friendly_error(exc))
            e.page.update()
        
        finally:
            _focus_first_field(e.page)
            try:
                e.page.update()
            except:
                pass

    def _handle_update(e: ft.ControlEvent) -> None:
        selected_id = state.get("selected_id")
        if not selected_id:
            _set_feedback("warning", "Güncelleme için listeden öğrenci seçiniz.")
            e.page.update()
            return

        validation_error = _validate_form()
        if validation_error:
            _set_feedback("warning", validation_error)
            e.page.update()
            return

        try:
            updated = controller.update_student(int(selected_id), _payload())
            if updated:
                _load_students()
                _refresh_form_from_selection()
                _render_table(e.page)
                _set_feedback("success", "Öğrenci başarıyla güncellendi.")
            else:
                _set_feedback("warning", "Güncellenecek kayıt bulunamadı.")
                _render_table(e.page)
        except Exception as exc:
            _set_feedback("error", _friendly_error(exc))
            _render_table(e.page)
        e.page.update()

    def _handle_select(record: dict, page: ft.Page | None) -> None:
        _fill_form(record)
        _set_feedback("info", "Öğrenci detayı forma yüklendi.")
        _render_table(page)
        if page:
            page.update()

    def _handle_delete(record: dict, page: ft.Page) -> None:
        record_id = int(record.get("id", 0) or 0)

        def _on_confirm(e: ft.ControlEvent) -> None:
            try:
                deleted = controller.delete_student(record_id)
                if deleted:
                    if int(state.get("selected_id") or 0) == record_id:
                        _clear_form()
                    _load_students()
                    _render_table(e.page)
                    _set_feedback("success", "Öğrenci başarıyla silindi.")
                    _focus_first_field(e.page)
                else:
                    _set_feedback("warning", "Silinecek kayıt bulunamadı.")
                    _render_table(e.page)
            except Exception as exc:
                _set_feedback("error", _friendly_error(exc))
                _render_table(e.page)

            e.page.close(dialog)
            e.page.update()

        def _on_cancel(e: ft.ControlEvent) -> None:
            _set_feedback("info", "Silme işlemi iptal edildi.")
            e.page.close(dialog)
            e.page.update()

        dialog = build_dialog(
            title="Silme Onayı",
            content=f"{record.get('ad_soyad') or 'Bu öğrenciyi'} silmek istediğinize emin misiniz?",
            confirm_text="Sil",
            cancel_text="Vazgeç",
            on_confirm=_on_confirm,
            on_cancel=_on_cancel,
        )
        page.open(dialog)
        page.update()

    def _previous_page(e: ft.ControlEvent) -> None:
        if state["page"] > 1:
            state["page"] -= 1
            _render_table(e.page)
            e.page.update()

    def _next_page(e: ft.ControlEvent, page_count: int) -> None:
        if state["page"] < page_count:
            state["page"] += 1
            _render_table(e.page)
            e.page.update()

    def _retry_state(e: ft.ControlEvent) -> None:
        _load_students()
        _render_table(e.page)
        e.page.update()

    def _render_table(page: ft.Page | None = None) -> None:
        if state.get("view_state") == "loading":
            table_host.controls = [build_loading_state("Ogrenci listesi yukleniyor")]
            return

        if state.get("view_state") == "error":
            table_host.controls = [build_error_state("Ogrenci listesi getirilemedi", on_retry=_retry_state)]
            return

        rows = list(state["students"])
        _sync_filter_state()
        filtered_rows = _students_v2_apply_filters(
            records=rows,
            search_query=state["search_query"],
            status_filter=state["status_filter"],
            class_filter=state["class_filter"],
            kur_filter=state["kur_filter"],
            start_from=state["start_from"],
            start_to=state["start_to"],
            sort_field=state["sort_field"],
            sort_direction=state["sort_direction"],
        )
        _render_active_filters()
        state["last_filtered_rows"] = list(filtered_rows)
        result_text.value = f"{len(rows)} ogrenciden {len(filtered_rows)} kayit gosteriliyor."

        if not rows:
            table_host.controls = [
                build_empty_state(
                    title="Henuz ogrenci bulunmuyor",
                    message="Yeni ogrenci ekleyerek baslayabilirsiniz.",
                    primary_action=build_primary_button("Yeni Ogrenci", icon=ft.Icons.PERSON_ADD, on_click=_handle_new),
                    secondary_action=build_secondary_button("Listeyi Yenile", on_click=_handle_list),
                )
            ]
            return

        if not filtered_rows:
            table_host.controls = [
                build_empty_state(
                    title="Kayit bulunamadi",
                    message="Arama kriterlerine uygun ogrenci bulunamadi.",
                    primary_action=build_secondary_button("Filtreleri Temizle", on_click=_handle_clear_filters),
                )
            ]
            return

        page_count = max(1, (len(filtered_rows) + state["page_size"] - 1) // state["page_size"])
        if state["page"] > page_count:
            state["page"] = page_count

        start = (state["page"] - 1) * state["page_size"]
        end = start + state["page_size"]
        page_rows = filtered_rows[start:end]

        selected_id = int(state.get("selected_id") or 0)
        table_rows: list[list[str | ft.Control]] = []

        for row in page_rows:
            row_id = int(row.get("id", 0) or 0)
            is_selected = row_id == selected_id
            row_bg = "#EEF2FF" if is_selected else "#FFFFFF"
            active_course_label = students_v2_active_course_label(row_id, state["courses"])

            ad_soyad_cell = ft.Container(
                bgcolor=row_bg,
                border_radius=8,
                padding=ft.Padding(8, 6, 8, 6),
                content=ft.TextButton(
                    content=ft.Text(
                        value=str(row.get("ad_soyad") or "-"),
                        weight=ft.FontWeight.W_600 if is_selected else ft.FontWeight.W_400,
                    ),
                    on_click=lambda e, r=row: _handle_select(r, e.page),
                    style=ft.ButtonStyle(
                        padding=ft.Padding(0, 0, 0, 0),
                        overlay_color=ft.Colors.TRANSPARENT,
                    ),
                ),
            )

            simple_cell = lambda value: ft.Container(
                bgcolor=row_bg,
                border_radius=8,
                padding=ft.Padding(8, 6, 8, 6),
                content=ft.Text(str(value)),
            )

            action_cell = ft.Container(
                bgcolor=row_bg,
                border_radius=8,
                padding=ft.Padding(6, 4, 6, 4),
                content=ft.Row(
                    spacing=8,
                    controls=[
                        build_secondary_button("Sec", on_click=lambda e, r=row: _handle_select(r, e.page)),
                        build_secondary_button("Kurs Ata", on_click=lambda e, r=row: _open_course_assignment_modal(r, e.page)),
                        build_danger_button("Sil", on_click=lambda e, r=row: _handle_delete(r, e.page)),
                    ],
                ),
            )

            table_rows.append(
                [
                    ad_soyad_cell,
                    simple_cell(str(row.get("sinif") or "-")),
                    simple_cell(_extract_kur(str(row.get("notlar") or "")) or "-"),
                    simple_cell(active_course_label),
                    simple_cell(str(row.get("telefon") or "-")),
                    ft.Container(
                        bgcolor=row_bg,
                        border_radius=8,
                        padding=ft.Padding(8, 6, 8, 6),
                        content=build_badge(
                            text=str(row.get("durum") or "Aktif"),
                            variant=_status_variant(str(row.get("durum") or "Aktif")),
                        ),
                    ),
                    simple_cell(str(row.get("baslangic_tarihi") or "-")),
                    action_cell,
                ]
            )

        pagination = ft.Row(
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            controls=[
                ft.Text(f"Toplam {len(filtered_rows)} kayit"),
                ft.Row(
                    spacing=8,
                    controls=[
                        build_secondary_button("Onceki", on_click=_previous_page, disabled=state["page"] <= 1),
                        ft.Text(f"Sayfa {state['page']} / {page_count}"),
                        build_secondary_button(
                            "Sonraki",
                            on_click=lambda e: _next_page(e, page_count),
                            disabled=state["page"] >= page_count,
                        ),
                    ],
                ),
            ],
        )

        table_host.controls = [
            build_table_card(
                title="Ogrenci Listesi",
                subtitle="Guncel ogrenci kayitlari",
                columns=["Ad Soyad", "Sinif", "Kur", "Aktif Kurs", "Telefon", "Durum", "Baslangic Tarihi", "Islemler"],
                rows=table_rows,
                footer=pagination,
            )
        ]

    header = build_app_header(
        title="Ogrenciler",
        subtitle="Ogrenci yonetim ekrani",
        leading=ft.Text("Ana Sayfa > Ogrenciler V2"),
        actions=[build_primary_button("Yeni Ogrenci", icon=ft.Icons.ADD, on_click=_handle_new)],
    )

    action_panel = build_action_panel(
        title="Hizli Islemler",
        subtitle="Import/Export islemleri",
        actions=[
            {
                "key": "new-student",
                "title": "Yeni Ogrenci",
                "subtitle": "Yeni kayit ac",
                "icon": ft.Icons.PERSON_ADD,
                "on_click": _handle_new,
            },
            {
                "key": "excel-export",
                "title": "Excel'e Aktar",
                "subtitle": "Filtreli listeyi XLSX indir",
                "icon": ft.Icons.DOWNLOAD,
                "on_click": lambda e: _handle_export_pick(e, "xlsx"),
            },
            {
                "key": "excel-import",
                "title": "Excel'den Ice Aktar",
                "subtitle": "XLSX dosyasindan toplu ekle",
                "icon": ft.Icons.UPLOAD,
                "on_click": _handle_import_pick,
            },
            {
                "key": "pdf-export",
                "title": "PDF'e Aktar",
                "subtitle": "Kurumsal rapor olarak indir",
                "icon": ft.Icons.PICTURE_AS_PDF,
                "on_click": lambda e: _handle_export_pick(e, "pdf"),
            },
            {
                "key": "csv-export",
                "title": "CSV'e Aktar",
                "subtitle": "Opsiyonel CSV cikti",
                "icon": ft.Icons.TABLE_VIEW,
                "on_click": lambda e: _handle_export_pick(e, "csv"),
            },
        ],
    )

    search_input_field: ft.TextField | None = None
    if isinstance(search_bar.content, ft.Column) and search_bar.content.controls:
        candidate = search_bar.content.controls[-1]
        if isinstance(candidate, ft.TextField):
            search_input_field = candidate

    if isinstance(search_input_field, ft.TextField):
        search_input_field.disabled = False
        search_input_field.on_change = lambda e: _handle_filter_change(e.page)

    filter_bar = build_filter_bar(
        title="Filtre Bar",
        subtitle="Durum, Sinif, Kur, Baslangic Tarihi ve Siralama",
        filters=[
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=filter_status),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=filter_class),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=filter_kur),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=sort_field_dropdown),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=sort_direction_dropdown),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=filter_start_from),
            ft.Container(col={"xs": 12, "sm": 6, "md": 3}, content=filter_start_to),
            ft.Container(
                col={"xs": 12, "sm": 6, "md": 3},
                content=build_secondary_button("Filtreleri Temizle", on_click=_handle_clear_filters),
            ),
        ],
    )

    result_card = build_app_card(
        title="Sonuc Bilgisi",
        subtitle="Filtre ve arama durumu",
        content=ft.Column(
            spacing=8,
            controls=[result_text, active_filters_row],
        ),
    )

    form_card = build_form_card(
        title="Yeni Ogrenci",
        subtitle="Ogrenci bilgilerini girin",
        fields=[
            ft.Container(col={"xs": 12, "md": 6}, content=name_field),
            ft.Container(col={"xs": 12, "md": 6}, content=class_field),
            ft.Container(col={"xs": 12, "md": 6}, content=parent_field),
            ft.Container(col={"xs": 12, "md": 6}, content=phone_field),
            ft.Container(col={"xs": 12, "md": 6}, content=email_field),
            ft.Container(col={"xs": 12, "md": 6}, content=username_field),
            ft.Container(col={"xs": 12, "md": 6}, content=password_field),
            ft.Container(col={"xs": 12, "md": 6}, content=start_date),
            ft.Container(col={"xs": 12, "md": 6}, content=end_date),
            ft.Container(col={"xs": 12, "md": 6}, content=kur_dropdown),
            ft.Container(col={"xs": 12, "md": 6}, content=status_dropdown),
            ft.Container(col={"xs": 12, "md": 12}, content=notes_field),
        ],
        actions=[
            build_secondary_button("Temizle", on_click=_handle_clear),
            build_secondary_button("Iptal", on_click=_handle_cancel),
            build_secondary_button("Listele", on_click=_handle_list),
            build_secondary_button("Guncelle", on_click=_handle_update),
            build_primary_button("Kaydet", on_click=_handle_create),
        ],
        max_height=450,  # Constrain form height for responsive layout
    )

    # Top section: header, search, filters (compact)
    top_section = ft.Column(
        spacing=16,
        controls=[
            header,
            action_panel,
            feedback_row,
            search_bar,
            filter_bar,
            result_card,
        ],
    )

    # Main section: form and table (scrollable independently)
    main_section = ft.Column(
        spacing=16,
        expand=True,
        controls=[
            form_card,
            table_host,
        ],
    )

    body = ft.Column(
        spacing=16,
        expand=True,
        controls=[
            top_section,
            main_section,
        ],
    )

    page = PageContainer(content=body, max_width=1888, padding=24)
    _load_students()
    _render_table()
    _set_feedback("info", "Hazır")
    return page
