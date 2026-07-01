"""Sprint S-6 Students V2 import/export validation tests."""

from __future__ import annotations

import tempfile
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT / "src") not in sys.path:
    sys.path.insert(0, str(ROOT / "src"))

from views.pages.students_v2 import (
    students_v2_export_excel_bytes,
    students_v2_export_pdf_bytes,
    students_v2_parse_import_xlsx,
)


def _build_valid_rows() -> list[dict]:
    return [
        {
            "ad_soyad": "Ali Veli",
            "sinif": "5-A",
            "telefon": "05551234567",
            "durum": "Aktif",
            "baslangic_tarihi": "2026-06-30",
            "notlar": "Kur: 2",
        },
        {
            "ad_soyad": "Ayse Kaya",
            "sinif": "6-B",
            "telefon": "05557654321",
            "durum": "Beklemede",
            "baslangic_tarihi": "2026-06-20",
            "notlar": "Not\nKur: 3",
        },
    ]


def test_excel_export() -> bool:
    data = students_v2_export_excel_bytes(_build_valid_rows())
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
        path.write_bytes(data)

    wb = load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    return header == ["Ad Soyad", "Sınıf", "Kur", "Telefon", "Durum", "Başlangıç"] and len(rows) == 2


def test_pdf_export() -> bool:
    data = students_v2_export_pdf_bytes(_build_valid_rows(), "Durum=Aktif")
    return data.startswith(b"%PDF") and len(data) > 300


def test_excel_import() -> bool:
    wb = Workbook()
    ws = wb.active
    ws.append(["Ad Soyad", "Sinif", "Kur", "Telefon", "E-posta", "Kullanici Adi", "Baslangic"])
    ws.append(["Ali Veli", "5-A", "2", "05551234567", "ali@example.com", "ali_user", "2026-06-30"])
    ws.append(["", "5-A", "2", "05551234567", "", "", "2026-06-30"])
    ws.append(["Mina", "", "", "abc", "not-mail", "ali_user", "2026-06-30"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
    wb.save(path)

    parsed = students_v2_parse_import_xlsx(str(path), existing_usernames=set())
    return parsed["total"] == 3 and len(parsed["valid_rows"]) == 1 and len(parsed["errors"]) == 2


def test_broken_file() -> bool:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, mode="w", encoding="utf-8") as handle:
        path = Path(handle.name)
        handle.write("bozuk dosya")

    try:
        students_v2_parse_import_xlsx(str(path), existing_usernames=set())
    except Exception:
        return True
    return False


def test_empty_file() -> bool:
    wb = Workbook()
    ws = wb.active
    ws.delete_rows(1, 1)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
    wb.save(path)

    parsed = students_v2_parse_import_xlsx(str(path), existing_usernames=set())
    return parsed["total"] == 0 and len(parsed["valid_rows"]) == 0 and len(parsed["errors"]) >= 1


def run_all() -> None:
    print("Excel Import Test", test_excel_import())
    print("Excel Export Test", test_excel_export())
    print("PDF Export Test", test_pdf_export())
    print("Bozuk Dosya Testi", test_broken_file())
    print("Bos Dosya Testi", test_empty_file())


if __name__ == "__main__":
    run_all()
